import pandas as pd
from openpyxl import load_workbook


def preprocess_excel(file_path: str) -> pd.DataFrame:
    """
    Preprocess an Excel file by:
    - Loading the first sheet
    - Handling merged cells by forward-filling values
    - Flattening multi-row headers into a single row
    - Stripping whitespace from column names
    - Dropping fully empty rows/columns
    
    Args:
        file_path: Path to the Excel file
        
    Returns:
        Clean 2D DataFrame suitable for analysis
    """
    # First, inspect the Excel file with openpyxl to detect multi-row headers
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    
    # Check first 3 rows to see if they look like multi-row headers
    # A header row typically has mostly text values (not numbers)
    header_rows = []
    for row_idx in range(1, min(4, ws.max_row + 1)):  # Check rows 1-3 (1-indexed)
        row_values = []
        text_count = 0
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value
            row_values.append(value)
            # Count text-like values (strings or non-numeric)
            if value is not None:
                if isinstance(value, str) or (isinstance(value, (int, float)) and row_idx == 1):
                    text_count += 1
        
        # If more than 50% of cells have values and they're mostly text, it's likely a header row
        non_null_count = sum(1 for v in row_values if v is not None)
        if non_null_count > 0 and (text_count / max(non_null_count, 1)) > 0.5:
            header_rows.append(row_idx - 1)  # Convert to 0-indexed
    
    # If we detected multiple header rows, flatten them
    if len(header_rows) > 1:
        # Collect header values from multiple rows
        header_data = []
        for row_idx in header_rows:
            row_values = []
            for col_idx in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_idx + 1, column=col_idx).value
                row_values.append(cell_value)
            header_data.append(row_values)
        
        # Flatten headers by joining with "_"
        flattened_headers = []
        max_cols = max(len(row) for row in header_data) if header_data else 0
        
        for col_idx in range(max_cols):
            header_parts = []
            for row in header_data:
                if col_idx < len(row) and row[col_idx] is not None:
                    header_parts.append(str(row[col_idx]).strip())
            flattened_header = '_'.join(header_parts).strip('_')
            flattened_headers.append(flattened_header if flattened_header else f'Column_{col_idx + 1}')
        
        # Read the dataframe, skipping the header rows
        df = pd.read_excel(file_path, sheet_name=0, engine='openpyxl', 
                          skiprows=max(header_rows) + 1, header=None)
        df.columns = flattened_headers[:len(df.columns)]
    else:
        # Single row header - read normally with pandas
        df = pd.read_excel(file_path, sheet_name=0, engine='openpyxl')
        # Strip whitespace from column names
        df.columns = [str(col).strip() if pd.notna(col) else f'Column_{i+1}' 
                     for i, col in enumerate(df.columns)]
    
    # Handle merged cells by forward-filling values
    # This fills NaN values with the previous non-null value in the same column
    df = df.ffill()
    
    # Strip whitespace from column names (ensure all are clean)
    df.columns = [str(col).strip() for col in df.columns]
    
    # Drop fully empty rows (all NaN)
    df = df.dropna(how='all')
    
    # Drop fully empty columns (all NaN)
    df = df.dropna(axis=1, how='all')
    
    # Reset index after dropping rows
    df = df.reset_index(drop=True)
    
    return df

